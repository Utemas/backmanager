from openpyxl import Workbook,load_workbook
from openpyxl.styles import *


import warnings
warnings.filterwarnings('ignore')

def findStartNum(ws):
    start_num = 1;
    for col in ws['A']:
        if col.value == '输入':
            return col.row + 1
def findEndNum(ws):
    end_num = 1;
    for col in ws['A']:
        if col.value == '输出':
            return col.row
seqNum = 23
phase1 = '<r-item itemname="'
phase2 = 'itemSeq="'
phase3 = 'label="'
phase4 = 'type="'
phase5 = 'length="'
phase6 = 'defaultvalue=""'
phase7 = 'isnull="'
phase8 = 'pkgitem="hashMap"/>'
loopstartend = 'pkgitem="hashMap" >'
phase9 = 'innerccode="'
rowList = []
totalList = []
mustFlag = ''
wb = load_workbook('./000.xlsx')
# openpyxl只能处理 .xlsx 合适的表格
ws = wb['PLMS0078']
start = findStartNum(ws)
end = findEndNum(ws)
for row in range(start,end,1):
    for column in ['H','I','J','K','L','M']:
        cell = ws[column+str(row)].value
        rowList.append(cell)
    
    totalList.append(rowList)
    rowList=[]
rowList = []
for outElemt in totalList:
    seqNum = seqNum+1
    if outElemt[4] == 'Y':
        mustFlag = 'false'
    else:
        mustFlag = 'true'
    str1 = phase1+outElemt[0]+'"'+'  '+phase2+str(seqNum)+'"'+'  '+phase3+outElemt[1]+'"'
    if '(' in outElemt[2]:
        if outElemt[2][:outElemt[2].index('(')] == 'Double':
            str1 = str1 +'  '+ phase4 + 'Number' + '"'
        else:
            str1 = str1 +'  '+ phase4 + outElemt[2][:outElemt[2].index('(')]+'"'
    else :
        if outElemt[5] == 'start':
            str1 = phase1 + outElemt[0] + '"' + '  ' + phase2 + str(seqNum) + '"' + '  ' + phase3 + outElemt[1]+'"' +'  '+ 'type="nodeloop"' + '  ' + phase5 +'8"' + '  ' + phase6 + '  '+ phase9 + outElemt[0] + '"' + '  ' + phase7 + mustFlag + '"' + '  ' + loopstartend
            print(str1)
            rowList.append(str1)
            continue
            print(seqNum)
        if outElemt[5] == 'end':
            str1 = '</r-item>'
            seqNum = seqNum -1
            print(str1)
            rowList.append(str1)
            continue
    if ')' in outElemt[2]:
        str1 = str1 + '  ' + phase5 + outElemt[2][outElemt[2].index('(')+1:outElemt[2].index(')')] +'"'
    str1 = str1 + '  ' + phase6 + '  '+ phase9 + outElemt[0] + '"' + '  ' + phase7 + mustFlag + '"' + '  ' +phase8
    print(str1)
    rowList.append(str1)
