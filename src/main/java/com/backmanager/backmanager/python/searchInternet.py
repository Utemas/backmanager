from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import warnings
warnings.filterwarnings('ignore')
wb = load_workbook('./000.xlsx')
ws = wb['PLMS0078']
end_num = 0
start_num = 1;
for col in ws['A']:
    if col.value == '输出':
        end_num = col.row - 1
        break
        # start_num = col.column
print(end_num)